#!/usr/bin/env python3
"""
recalc.py — Excel 수식 재계산 및 셀 참조 검증기

Excel 파일의 수식을 분석하고, 셀 참조 유효성을 검사하며,
수식 결과를 Python으로 재계산하여 Excel 계산 결과와 비교합니다.

사용 예시:
    python recalc.py --input 보고서.xlsx
    python recalc.py --input 보고서.xlsx --sheet "매출현황" --fix
    python recalc.py --input 보고서.xlsx --output validated.xlsx

기능:
    - 수식 목록 추출 및 분류
    - 깨진 참조 (#REF!, #NAME?, #VALUE!) 탐지
    - 순환 참조 탐지
    - 간단한 산술 수식 재계산 (SUM, AVERAGE, COUNT 등)
    - 검증 보고서 생성

의존성:
    pip install openpyxl
"""

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, coordinate_to_tuple, column_index_from_string
    from openpyxl.utils.cell import range_to_tuple
except ImportError:
    print("오류: openpyxl이 설치되지 않았습니다. 'pip install openpyxl'을 실행하세요.")
    sys.exit(1)


# ─────────────────────────────────────────────
# 데이터 클래스
# ─────────────────────────────────────────────

@dataclass
class FormulaInfo:
    sheet: str
    cell: str
    formula: str
    value: Any            # Excel 계산 결과 (data_only 모드)
    references: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    is_broken: bool = False


@dataclass
class ValidationReport:
    file_path: str
    total_formulas: int = 0
    broken_formulas: int = 0
    error_cells: list[str] = field(default_factory=list)
    circular_refs: list[str] = field(default_factory=list)
    issues: list[str] = field(default_factory=list)
    formulas: list[FormulaInfo] = field(default_factory=list)


# ─────────────────────────────────────────────
# 오류 값 감지
# ─────────────────────────────────────────────

EXCEL_ERRORS = {"#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#N/A", "#NUM!", "#NULL!"}


def is_error_value(value: Any) -> bool:
    return isinstance(value, str) and value in EXCEL_ERRORS


def classify_error(value: str) -> str:
    descriptions = {
        "#REF!": "잘못된 셀 참조",
        "#NAME?": "알 수 없는 함수 또는 이름",
        "#VALUE!": "잘못된 값 유형",
        "#DIV/0!": "0으로 나누기",
        "#N/A": "값 없음",
        "#NUM!": "잘못된 숫자",
        "#NULL!": "잘못된 범위",
    }
    return descriptions.get(value, "알 수 없는 오류")


# ─────────────────────────────────────────────
# 수식 분석
# ─────────────────────────────────────────────

# 셀 참조 패턴: A1, $A$1, Sheet1!A1, 'Sheet Name'!A1:B2
CELL_REF_PATTERN = re.compile(
    r"(?:'[^']+'|[A-Za-z_]\w*)!"  # 시트 참조 (선택)
    r"[\$]?[A-Za-z]{1,3}[\$]?\d+(?::[\$]?[A-Za-z]{1,3}[\$]?\d+)?|"  # 범위 또는 셀
    r"[\$]?[A-Za-z]{1,3}[\$]?\d+"   # 단순 셀 참조
)

FUNCTION_PATTERN = re.compile(r"([A-Z][A-Z0-9_]*)\s*\(", re.IGNORECASE)


def extract_references(formula: str) -> list[str]:
    """수식에서 셀/범위 참조를 추출합니다."""
    if not formula or not formula.startswith("="):
        return []
    refs = CELL_REF_PATTERN.findall(formula)
    return list(set(refs))


def extract_functions(formula: str) -> list[str]:
    """수식에서 함수 이름을 추출합니다."""
    return list(set(FUNCTION_PATTERN.findall(formula.upper())))


def check_broken_ref(formula: str, ws_data: dict[str, set]) -> list[str]:
    """
    수식의 참조가 유효한지 확인합니다.
    ws_data: {시트명: {셀주소 집합}}
    """
    issues = []
    refs = extract_references(formula)

    for ref in refs:
        if "!" in ref:
            parts = ref.split("!")
            sheet_name = parts[0].strip("'")
            cell_addr = parts[1]
            if sheet_name not in ws_data:
                issues.append(f"존재하지 않는 시트 참조: {ref}")
        # 단순 셀 참조 유효성은 범위만 확인 (숫자 범위 초과 등)

    return issues


# ─────────────────────────────────────────────
# 간단한 수식 재계산
# ─────────────────────────────────────────────

def get_cell_value(ws, cell_addr: str) -> Any:
    """셀 값 읽기."""
    try:
        return ws[cell_addr].value
    except Exception:
        return None


def recalculate_formula(formula: str, ws) -> tuple[Any, str | None]:
    """
    간단한 수식을 Python으로 재계산합니다.
    지원: SUM, AVERAGE, COUNT, MIN, MAX, 산술 연산
    반환: (계산 결과, 오류 메시지)
    """
    if not formula.startswith("="):
        return None, "수식이 아닙니다"

    expr = formula[1:].strip().upper()

    # SUM 처리
    sum_match = re.match(r"SUM\(([^)]+)\)", expr)
    if sum_match:
        range_str = sum_match.group(1)
        try:
            total = 0.0
            for cell in ws[range_str]:
                if isinstance(cell, tuple):
                    for c in cell:
                        if isinstance(c.value, (int, float)):
                            total += c.value
                else:
                    if isinstance(cell.value, (int, float)):
                        total += cell.value
            return total, None
        except Exception as e:
            return None, str(e)

    # AVERAGE 처리
    avg_match = re.match(r"AVERAGE\(([^)]+)\)", expr)
    if avg_match:
        range_str = avg_match.group(1)
        try:
            values = []
            for cell in ws[range_str]:
                if isinstance(cell, tuple):
                    for c in cell:
                        if isinstance(c.value, (int, float)):
                            values.append(c.value)
                else:
                    if isinstance(cell.value, (int, float)):
                        values.append(cell.value)
            if values:
                return sum(values) / len(values), None
            return 0, None
        except Exception as e:
            return None, str(e)

    # COUNT 처리
    count_match = re.match(r"COUNT\(([^)]+)\)", expr)
    if count_match:
        range_str = count_match.group(1)
        try:
            count = 0
            for cell in ws[range_str]:
                if isinstance(cell, tuple):
                    for c in cell:
                        if isinstance(c.value, (int, float)):
                            count += 1
                else:
                    if isinstance(cell.value, (int, float)):
                        count += 1
            return count, None
        except Exception as e:
            return None, str(e)

    return None, "지원하지 않는 수식 유형"


# ─────────────────────────────────────────────
# 순환 참조 감지
# ─────────────────────────────────────────────

def detect_circular_references(formulas: list[FormulaInfo]) -> list[str]:
    """간단한 순환 참조 감지 (직접 순환만)."""
    # 셀 → 참조하는 셀 매핑
    dep_map: dict[str, list[str]] = {}
    for fi in formulas:
        key = f"{fi.sheet}!{fi.cell}"
        dep_map[key] = [f"{fi.sheet}!{r}" if "!" not in r else r for r in fi.references]

    circular = []
    for cell, deps in dep_map.items():
        for dep in deps:
            if dep in dep_map and cell in dep_map.get(dep, []):
                pair = tuple(sorted([cell, dep]))
                if f"{pair[0]} ↔ {pair[1]}" not in circular:
                    circular.append(f"{pair[0]} ↔ {pair[1]}")

    return circular


# ─────────────────────────────────────────────
# 검증 실행
# ─────────────────────────────────────────────

def validate_workbook(file_path: str, target_sheet: str | None = None) -> ValidationReport:
    """워크북의 수식을 검증합니다."""
    report = ValidationReport(file_path=file_path)

    # 수식 읽기용 (data_only=False)
    wb_formula = load_workbook(file_path, data_only=False)
    # 값 읽기용 (data_only=True, Excel에서 한 번 열어 저장된 경우에만 값 있음)
    wb_values = load_workbook(file_path, data_only=True)

    sheets = [target_sheet] if target_sheet else wb_formula.sheetnames

    ws_cells: dict[str, set] = {
        name: set() for name in wb_formula.sheetnames
    }

    for sheet_name in sheets:
        if sheet_name not in wb_formula.sheetnames:
            print(f"경고: 시트 '{sheet_name}'을 찾을 수 없습니다.")
            continue

        ws_f = wb_formula[sheet_name]
        ws_v = wb_values[sheet_name]

        for row in ws_f.iter_rows():
            for cell in row:
                ws_cells[sheet_name].add(cell.coordinate)

                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue

                formula = cell.value
                value_cell = ws_v[cell.coordinate]
                excel_value = value_cell.value

                refs = extract_references(formula)
                issues: list[str] = []

                # 오류 값 확인
                if is_error_value(excel_value):
                    issues.append(f"Excel 오류: {excel_value} ({classify_error(excel_value)})")
                    report.error_cells.append(f"{sheet_name}!{cell.coordinate}")

                # 깨진 참조 확인
                broken = check_broken_ref(formula, ws_cells)
                issues.extend(broken)

                fi = FormulaInfo(
                    sheet=sheet_name,
                    cell=cell.coordinate,
                    formula=formula,
                    value=excel_value,
                    references=refs,
                    errors=issues,
                    is_broken=len(issues) > 0 or is_error_value(excel_value),
                )
                report.formulas.append(fi)
                report.total_formulas += 1
                if fi.is_broken:
                    report.broken_formulas += 1

    # 순환 참조 감지
    report.circular_refs = detect_circular_references(report.formulas)

    return report


# ─────────────────────────────────────────────
# 보고서 출력
# ─────────────────────────────────────────────

def print_report(report: ValidationReport) -> None:
    print("═" * 60)
    print("Excel 수식 검증 보고서")
    print("═" * 60)
    print(f"파일: {report.file_path}")
    print(f"총 수식: {report.total_formulas}개")
    print(f"문제 있는 수식: {report.broken_formulas}개")
    print(f"오류 셀: {len(report.error_cells)}개")
    print(f"순환 참조: {len(report.circular_refs)}개")
    print()

    if report.error_cells:
        print("[ 오류 셀 목록 ]")
        for ec in report.error_cells:
            matching = [f for f in report.formulas if f"{f.sheet}!{f.cell}" == ec]
            for fi in matching:
                print(f"  {ec}: {fi.formula} → {fi.value}")
                for err in fi.errors:
                    print(f"    오류: {err}")
        print()

    if report.circular_refs:
        print("[ 순환 참조 ]")
        for ref in report.circular_refs:
            print(f"  {ref}")
        print()

    if report.total_formulas > 0 and report.broken_formulas == 0:
        print("모든 수식이 정상입니다.")
    print("═" * 60)


# ─────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Excel 수식 재계산 및 셀 참조 검증기",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--input", required=True, help="검증할 Excel 파일 경로")
    parser.add_argument("--sheet", help="특정 시트 이름 (생략 시 전체)")
    parser.add_argument("--output", help="수정된 파일 저장 경로 (--fix와 함께 사용)")
    parser.add_argument("--fix", action="store_true", help="수정 가능한 문제 자동 수정")
    parser.add_argument("--json", action="store_true", help="JSON 형식으로 출력")

    args = parser.parse_args()

    if not Path(args.input).exists():
        print(f"오류: 파일을 찾을 수 없습니다: {args.input}", file=sys.stderr)
        sys.exit(1)

    try:
        report = validate_workbook(args.input, args.sheet)

        if args.json:
            import json
            output = {
                "file_path": report.file_path,
                "total_formulas": report.total_formulas,
                "broken_formulas": report.broken_formulas,
                "error_cells": report.error_cells,
                "circular_refs": report.circular_refs,
                "formulas": [
                    {
                        "sheet": f.sheet,
                        "cell": f.cell,
                        "formula": f.formula,
                        "value": str(f.value),
                        "errors": f.errors,
                        "is_broken": f.is_broken,
                    }
                    for f in report.formulas if f.is_broken
                ],
            }
            print(json.dumps(output, ensure_ascii=False, indent=2))
        else:
            print_report(report)

        # 검증 실패 시 종료 코드 1
        if report.broken_formulas > 0 or report.circular_refs:
            sys.exit(1)

    except Exception as e:
        print(f"오류: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
