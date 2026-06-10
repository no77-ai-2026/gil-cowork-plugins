#!/usr/bin/env python3
"""
read_xlsx.py — Excel 파일 분석기 및 요약기

Excel 파일의 시트, 컬럼, 데이터 유형, 통계를 분석하여 요약 보고서를 출력합니다.
한국어 텍스트 및 원화 통화 형식을 지원합니다.

사용 예시:
    python read_xlsx.py --input 데이터.xlsx
    python read_xlsx.py --input 데이터.xlsx --sheet "매출현황"
    python read_xlsx.py --input 데이터.xlsx --output summary.json
    python read_xlsx.py --input 데이터.xlsx --rows 5  # 미리보기 행 수

의존성:
    pip install openpyxl
"""

import argparse
import json
import sys
from collections import Counter
from datetime import datetime, date
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("오류: openpyxl이 설치되지 않았습니다. 'pip install openpyxl'을 실행하세요.")
    sys.exit(1)


# ─────────────────────────────────────────────
# 데이터 유형 감지
# ─────────────────────────────────────────────

def detect_type(value: Any) -> str:
    """셀 값의 데이터 유형을 감지합니다."""
    if value is None:
        return "empty"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "float"
    if isinstance(value, (datetime, date)):
        return "datetime"
    if isinstance(value, str):
        if not value.strip():
            return "empty"
        # 한국어 포함 여부
        has_korean = any('\uAC00' <= c <= '\uD7A3' for c in value)
        return "korean_text" if has_korean else "text"
    return "unknown"


def is_numeric(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


# ─────────────────────────────────────────────
# 컬럼 분석
# ─────────────────────────────────────────────

def analyze_column(values: list[Any]) -> dict[str, Any]:
    """단일 컬럼을 분석합니다."""
    non_null = [v for v in values if v is not None and v != ""]
    type_counts = Counter(detect_type(v) for v in values)
    numeric_vals = [v for v in non_null if is_numeric(v)]

    stats: dict[str, Any] = {
        "total_rows": len(values),
        "non_null": len(non_null),
        "null_count": len(values) - len(non_null),
        "null_ratio": round((len(values) - len(non_null)) / max(len(values), 1), 3),
        "unique_count": len(set(str(v) for v in non_null)),
        "dominant_type": type_counts.most_common(1)[0][0] if type_counts else "empty",
        "type_distribution": dict(type_counts),
    }

    # 숫자형 통계
    if numeric_vals:
        stats["numeric_stats"] = {
            "min": min(numeric_vals),
            "max": max(numeric_vals),
            "sum": sum(numeric_vals),
            "mean": round(sum(numeric_vals) / len(numeric_vals), 2),
        }

    # 텍스트 길이 통계
    text_vals = [v for v in non_null if isinstance(v, str)]
    if text_vals:
        lengths = [len(v) for v in text_vals]
        stats["text_stats"] = {
            "avg_length": round(sum(lengths) / len(lengths), 1),
            "max_length": max(lengths),
            "sample_values": list(set(text_vals))[:5],
        }

    return stats


# ─────────────────────────────────────────────
# 시트 분석
# ─────────────────────────────────────────────

def analyze_sheet(ws) -> dict[str, Any]:
    """워크시트를 분석합니다."""
    # 헤더 감지 (첫 번째 행)
    headers: list[str] = []
    if ws.max_row > 0:
        first_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        # 첫 행이 텍스트이면 헤더로 간주
        if all(isinstance(v, str) or v is None for v in first_row):
            headers = [str(v) if v is not None else f"컬럼{i+1}" for i, v in enumerate(first_row)]
            data_start_row = 2
        else:
            headers = [f"컬럼{i+1}" for i in range(ws.max_column)]
            data_start_row = 1
    else:
        data_start_row = 1

    # 데이터 읽기
    columns_data: list[list[Any]] = [[] for _ in range(ws.max_column)]
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        for col_idx, value in enumerate(row):
            if col_idx < len(columns_data):
                columns_data[col_idx].append(value)

    # 컬럼별 분석
    columns_info: list[dict] = []
    for idx, (header, col_data) in enumerate(zip(headers, columns_data)):
        col_analysis = analyze_column(col_data)
        col_analysis["name"] = header
        col_analysis["column_letter"] = get_column_letter(idx + 1)
        columns_info.append(col_analysis)

    # 병합 셀 확인
    merged_ranges = [str(m) for m in ws.merged_cells.ranges]

    # 수식 셀 수
    formula_count = sum(
        1 for row in ws.iter_rows(values_only=False)
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )

    return {
        "name": ws.title,
        "dimensions": {
            "rows": ws.max_row,
            "columns": ws.max_column,
            "data_rows": ws.max_row - data_start_row + 1,
        },
        "has_header": data_start_row == 2,
        "merged_ranges": merged_ranges,
        "formula_count": formula_count,
        "columns": columns_info,
    }


# ─────────────────────────────────────────────
# 미리보기 생성
# ─────────────────────────────────────────────

def get_preview(ws, num_rows: int = 5) -> list[list[str]]:
    """시트의 처음 N행을 미리보기로 반환합니다."""
    preview = []
    for row in ws.iter_rows(max_row=num_rows + 1, values_only=True):
        preview.append([str(v) if v is not None else "" for v in row])
    return preview


# ─────────────────────────────────────────────
# 포맷 출력
# ─────────────────────────────────────────────

def format_text_report(file_path: str, wb, sheet_analyses: list[dict], preview_rows: int) -> str:
    lines = []
    lines.append("═" * 60)
    lines.append("Excel 파일 분석 보고서")
    lines.append("═" * 60)
    lines.append(f"파일: {file_path}")
    lines.append(f"분석 시각: {datetime.now().strftime('%Y년 %m월 %d일 %H:%M:%S')}")
    lines.append(f"시트 수: {len(wb.sheetnames)}개")
    lines.append(f"시트 목록: {', '.join(wb.sheetnames)}")
    lines.append("")

    for analysis in sheet_analyses:
        lines.append(f"[ 시트: {analysis['name']} ]")
        dim = analysis["dimensions"]
        lines.append(f"  크기: {dim['rows']}행 × {dim['columns']}열 (데이터: {dim['data_rows']}행)")
        lines.append(f"  헤더 행 존재: {'예' if analysis['has_header'] else '아니오'}")
        if analysis["formula_count"] > 0:
            lines.append(f"  수식 셀: {analysis['formula_count']}개")
        if analysis["merged_ranges"]:
            lines.append(f"  병합 셀: {len(analysis['merged_ranges'])}개")

        lines.append("\n  [ 컬럼 분석 ]")
        for col in analysis["columns"]:
            type_info = col["dominant_type"]
            null_info = f"누락 {col['null_count']}개" if col["null_count"] > 0 else "누락 없음"
            lines.append(f"  {col['column_letter']:>3}. {col['name']:<20} | 유형: {type_info:<15} | {null_info}")

            if "numeric_stats" in col:
                ns = col["numeric_stats"]
                lines.append(f"       숫자 통계: 최소={ns['min']:,}, 최대={ns['max']:,}, 합계={ns['sum']:,.0f}, 평균={ns['mean']:,.1f}")

            if "text_stats" in col and col["text_stats"]["sample_values"]:
                samples = col["text_stats"]["sample_values"][:3]
                lines.append(f"       샘플 값: {', '.join(samples)}")

        lines.append("")

    return "\n".join(lines)


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Excel 파일 분석기 및 요약기",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--input", required=True, help="분석할 Excel 파일 경로")
    parser.add_argument("--sheet", help="분석할 특정 시트 이름 (생략 시 전체 시트)")
    parser.add_argument("--output", help="결과 저장 경로 (.json)")
    parser.add_argument("--rows", type=int, default=5, help="미리보기 행 수 (기본: 5)")
    parser.add_argument("--format", choices=["text", "json"], default="text", help="출력 형식")

    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"오류: 파일을 찾을 수 없습니다: {args.input}", file=sys.stderr)
        sys.exit(1)

    try:
        # data_only=True: 수식 결과값 읽기
        wb = load_workbook(str(input_path), data_only=True)

        target_sheets = [args.sheet] if args.sheet else wb.sheetnames

        sheet_analyses = []
        for sheet_name in target_sheets:
            if sheet_name not in wb.sheetnames:
                print(f"경고: 시트 '{sheet_name}'을 찾을 수 없습니다.", file=sys.stderr)
                continue
            ws = wb[sheet_name]
            analysis = analyze_sheet(ws)
            if args.rows > 0:
                analysis["preview"] = get_preview(ws, args.rows)
            sheet_analyses.append(analysis)

        result = {
            "file_path": str(input_path),
            "analyzed_at": datetime.now().isoformat(),
            "file_size_kb": round(input_path.stat().st_size / 1024, 1),
            "sheets": sheet_analyses,
        }

        if args.format == "json" or args.output:
            output_text = json.dumps(result, ensure_ascii=False, indent=2, default=str)
        else:
            output_text = format_text_report(str(input_path), wb, sheet_analyses, args.rows)

        if args.output:
            Path(args.output).write_text(output_text, encoding="utf-8")
            print(f"결과 저장: {args.output}")
        else:
            print(output_text)

    except Exception as e:
        print(f"오류: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
