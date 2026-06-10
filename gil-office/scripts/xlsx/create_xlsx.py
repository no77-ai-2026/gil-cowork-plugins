#!/usr/bin/env python3
"""
create_xlsx.py — Excel 워크북 생성기 (한국어 포맷 지원)

openpyxl을 사용하여 한국어 비즈니스 문서에 적합한 Excel 워크북을 생성합니다.
원화(₩) 통화 형식, 한국식 날짜 형식, 한국어 폰트를 지원합니다.

사용 예시:
    python create_xlsx.py --output 보고서.xlsx --title "2026년 Q1 실적" --spec spec.json
    python create_xlsx.py --output 견적서.xlsx --template invoice

의존성:
    pip install openpyxl

spec.json 구조:
{
    "sheets": [
        {
            "name": "매출 현황",
            "headers": ["날짜", "품목", "수량", "단가", "합계"],
            "rows": [
                ["2026-04-01", "서비스A", 10, 150000, "=C2*D2"],
                ["2026-04-02", "서비스B", 5, 200000, "=C3*D3"]
            ],
            "formats": {
                "A": "date",
                "D": "currency",
                "E": "currency"
            }
        }
    ]
}
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Fill, PatternFill, Alignment, Border, Side,
        numbers
    )
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print("오류: openpyxl이 설치되지 않았습니다. 'pip install openpyxl'을 실행하세요.")
    sys.exit(1)


# ─────────────────────────────────────────────
# 한국어 형식 상수
# ─────────────────────────────────────────────

KO_DATE_FORMAT = "YYYY년 MM월 DD일"
KO_DATETIME_FORMAT = "YYYY년 MM월 DD일 HH시 MM분"
KO_CURRENCY_FORMAT = '₩#,##0'
KO_CURRENCY_FORMAT_DECIMAL = '₩#,##0.00'
KO_NUMBER_FORMAT = '#,##0'
KO_PERCENT_FORMAT = '0.0%'
KO_FONT = "맑은 고딕"

# 형식 코드 매핑
FORMAT_MAP = {
    "date":       KO_DATE_FORMAT,
    "datetime":   KO_DATETIME_FORMAT,
    "currency":   KO_CURRENCY_FORMAT,
    "currency_d": KO_CURRENCY_FORMAT_DECIMAL,
    "number":     KO_NUMBER_FORMAT,
    "percent":    KO_PERCENT_FORMAT,
    "text":       "@",
    "general":    "General",
}


# ─────────────────────────────────────────────
# 스타일 헬퍼
# ─────────────────────────────────────────────

def header_font(size: int = 11) -> Font:
    """헤더 셀 폰트."""
    return Font(name=KO_FONT, bold=True, size=size, color="FFFFFF")


def body_font(size: int = 10) -> Font:
    """본문 셀 폰트."""
    return Font(name=KO_FONT, size=size)


def title_font(size: int = 16) -> Font:
    """제목 폰트."""
    return Font(name=KO_FONT, bold=True, size=size, color="1F3864")


def header_fill(color: str = "2E5090") -> PatternFill:
    """헤더 배경색."""
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def alt_row_fill(color: str = "EEF2F7") -> PatternFill:
    """교대 행 배경색."""
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def thin_border() -> Border:
    """얇은 테두리."""
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def center_align() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def right_align() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


def left_align() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ─────────────────────────────────────────────
# 시트 생성
# ─────────────────────────────────────────────

def auto_fit_columns(ws: Worksheet, min_width: int = 8, max_width: int = 50) -> None:
    """열 너비 자동 조정."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                # 한국어 문자는 2배 너비로 계산
                text = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in text)
                max_length = max(max_length, length)
        adjusted = min(max_width, max(min_width, max_length + 2))
        ws.column_dimensions[col_letter].width = adjusted


def create_sheet_from_spec(wb: Workbook, sheet_spec: dict[str, Any]) -> Worksheet:
    """스펙에서 워크시트를 생성합니다."""
    name = sheet_spec.get("name", "시트1")
    ws = wb.create_sheet(title=name)

    headers: list[str] = sheet_spec.get("headers", [])
    rows: list[list] = sheet_spec.get("rows", [])
    formats: dict[str, str] = sheet_spec.get("formats", {})
    freeze_row: int = sheet_spec.get("freeze_row", 1)
    title_text: str = sheet_spec.get("title", "")
    row_offset = 0

    # 제목 행 추가
    if title_text:
        ws.cell(row=1, column=1, value=title_text).font = title_font()
        if headers:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=1).alignment = center_align()
        ws.row_dimensions[1].height = 30
        row_offset = 1

    # 헤더 행
    if headers:
        header_row = row_offset + 1
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font()
            cell.fill = header_fill()
            cell.alignment = center_align()
            cell.border = thin_border()
        ws.row_dimensions[header_row].height = 20
        row_offset += 1

    # 데이터 행
    for row_idx, row_data in enumerate(rows, 1):
        actual_row = row_offset + row_idx
        use_alt_fill = row_idx % 2 == 0

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=actual_row, column=col_idx, value=value)
            cell.font = body_font()
            cell.border = thin_border()

            if use_alt_fill:
                cell.fill = alt_row_fill()

            # 형식 적용
            col_letter = get_column_letter(col_idx)
            fmt_key = formats.get(col_letter, "")
            if fmt_key and fmt_key in FORMAT_MAP:
                cell.number_format = FORMAT_MAP[fmt_key]
                if fmt_key in ("currency", "currency_d", "number"):
                    cell.alignment = right_align()
                elif fmt_key == "date":
                    cell.alignment = center_align()
                else:
                    cell.alignment = left_align()
            else:
                if isinstance(value, (int, float)):
                    cell.alignment = right_align()
                else:
                    cell.alignment = left_align()

        ws.row_dimensions[actual_row].height = 18

    # 열 자동 맞춤
    auto_fit_columns(ws)

    # 틀 고정
    if freeze_row > 0 and headers:
        ws.freeze_panes = ws.cell(row=freeze_row + row_offset + 1, column=1)

    return ws


# ─────────────────────────────────────────────
# 내장 템플릿
# ─────────────────────────────────────────────

BUILTIN_TEMPLATES: dict[str, dict] = {
    "invoice": {
        "sheets": [
            {
                "name": "견적서",
                "title": "견 적 서",
                "headers": ["번호", "품목", "규격", "수량", "단가(₩)", "공급가액(₩)", "부가세(₩)", "합계(₩)"],
                "rows": [
                    [1, "서비스명", "표준형", 1, 500000, "=E2*D2", "=F2*0.1", "=F2+G2"],
                    [2, "추가 서비스", "기본형", 2, 150000, "=E3*D3", "=F3*0.1", "=F3+G3"],
                ],
                "formats": {
                    "E": "currency", "F": "currency", "G": "currency", "H": "currency"
                },
            }
        ]
    },
    "monthly_report": {
        "sheets": [
            {
                "name": "월간현황",
                "title": f"{datetime.now().year}년 {datetime.now().month}월 현황 보고",
                "headers": ["날짜", "항목", "구분", "금액(₩)", "비고"],
                "rows": [],
                "formats": {"A": "date", "D": "currency"},
            }
        ]
    },
    "budget": {
        "sheets": [
            {
                "name": "예산안",
                "title": "부서별 예산안",
                "headers": ["부서", "항목", "예산(₩)", "집행(₩)", "잔액(₩)", "집행률(%)"],
                "rows": [
                    ["개발팀", "인건비", 10000000, 7500000, "=C2-D2", "=D2/C2"],
                    ["마케팅팀", "광고비", 5000000, 3200000, "=C3-D3", "=D3/C3"],
                ],
                "formats": {"C": "currency", "D": "currency", "E": "currency", "F": "percent"},
            }
        ]
    },
}


# ─────────────────────────────────────────────
# 메인 생성 함수
# ─────────────────────────────────────────────

def create_workbook(
    output_path: str,
    spec: dict[str, Any] | None = None,
    template: str | None = None,
    title: str | None = None,
) -> None:
    """Excel 워크북 생성."""
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    # 스펙 결정
    if template and template in BUILTIN_TEMPLATES:
        final_spec = BUILTIN_TEMPLATES[template]
    elif spec:
        final_spec = spec
    else:
        # 기본 빈 시트
        final_spec = {
            "sheets": [{"name": "데이터", "headers": ["항목", "값", "비고"], "rows": []}]
        }

    # 문서 속성
    wb.properties.title = title or final_spec.get("title", "")
    wb.properties.creator = "MoAI Office Plugin"
    wb.properties.created = datetime.now()
    wb.properties.modified = datetime.now()

    # 시트 생성
    for sheet_spec in final_spec.get("sheets", []):
        create_sheet_from_spec(wb, sheet_spec)

    if not wb.sheetnames:
        wb.create_sheet("데이터")

    # 저장
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    print(f"저장 완료: {output} ({len(wb.sheetnames)}시트)")


# ─────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Excel 워크북 생성기 (한국어 포맷 지원)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--output", required=True, help="출력 Excel 파일 경로 (.xlsx)")
    parser.add_argument("--spec", help="스펙 JSON 파일 경로")
    parser.add_argument("--template", choices=list(BUILTIN_TEMPLATES.keys()), help="내장 템플릿 사용")
    parser.add_argument("--title", help="워크북 제목")

    args = parser.parse_args()

    spec = None
    if args.spec:
        with open(args.spec, encoding="utf-8") as f:
            spec = json.load(f)

    try:
        create_workbook(
            output_path=args.output,
            spec=spec,
            template=args.template,
            title=args.title,
        )
    except Exception as e:
        print(f"오류: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
