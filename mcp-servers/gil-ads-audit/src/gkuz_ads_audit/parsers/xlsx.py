"""
.xlsx 보고서 파서 — openpyxl 단독, PIPA 메모리만 처리.

REQ-AUDIT-MCP-016: 12개 필수 컬럼 자동 매핑 + 정합성 검증 6종 + 누락 처리 4룰.
REQ-AUDIT-MCP-019: .xlsx 원본·중간 가공물 디스크 저장 금지 (메모리만 처리, PIPA).
REQ-AUDIT-MCP-005: xlsx 경로 fallback 지원.

주의:
    - openpyxl 이외 xlsx 라이브러리 (pandas, xlrd 등) 의존성 금지.
    - 자격증명·인구통계 raw 데이터를 로그·stdout에 노출 금지 (REQ-AUDIT-MCP-023).
"""
from __future__ import annotations

import io
import os
from dataclasses import dataclass, field
from typing import Any

import openpyxl
from openpyxl.workbook import Workbook


# ============================================================
# 12 필수 컬럼 매핑 — SPEC §3.4 한국어↔영어 양방향
# ============================================================

# @MX:ANCHOR: [AUTO] 12 필수 컬럼 SSOT — xlsx.py·tests·audit_meta_account.py 공통 참조
# @MX:REASON: SPEC §3.4 "12 필수 컬럼 자동 매핑" REQ-AUDIT-MCP-016 검증 조건. 컬럼 추가·삭제 시 tests 재검증 필수.
REQUIRED_COLUMN_ALIASES: dict[str, list[str]] = {
    # 표준 키: [한국어 별칭, 영어 별칭, ...]
    "reporting_starts": ["보고 시작", "Reporting starts", "Start date"],
    "reporting_ends": ["보고 종료", "Reporting ends", "End date"],
    "ad_name": ["광고 이름", "Ad name", "Ad Name"],
    "placement": ["노출 위치", "Placement"],
    "age": ["연령", "Age"],
    "gender": ["성별", "Gender"],
    "amount_spent": ["지출 금액", "Amount spent", "Spend"],
    "link_clicks": ["링크 클릭", "Link clicks", "Clicks"],
    "ctr": ["CTR", "CTR (link CTR)", "Link CTR"],
    "cpc": ["CPC", "CPC (link CPC)", "Link CPC"],
    "purchases": ["구매", "Purchases", "Purchase"],
    "purchase_conversion_value": ["구매 전환값", "Purchase conversion value", "Revenue"],
    "purchase_roas": ["구매 ROAS", "Purchase ROAS", "ROAS"],
}

# 필수 컬럼 키 집합 (연령·성별은 누락 허용, 인구통계 분석 비활성)
MANDATORY_COLUMNS = {
    "reporting_starts", "reporting_ends", "ad_name",
    "amount_spent", "purchases", "purchase_conversion_value", "purchase_roas",
}

# 선택 컬럼 (누락 시 관련 기능 비활성)
OPTIONAL_COLUMNS = {"placement", "age", "gender", "link_clicks", "ctr", "cpc"}

# 표본 부족 셀 마스킹 N값 — SPEC §7.2 결정 (v0.2.0, 2026-05-13)
# 사유: 한국 PIPA 재식별 위험 + 통계적 유의성. 인구통계 cross-tab 셀 카운트 < 5 시 마스킹.
# @MX:ANCHOR: [AUTO] 표본 부족 마스킹 N값 SSOT — xlsx.py·tests·downstream audit 도구 공통 참조
# @MX:REASON: SPEC §7.2 결정 — PIPA 개인정보 보호 + 통계 유의성. M-CR3 임계값 1000과는 별개.
MASKING_CELL_THRESHOLD = 5


# ============================================================
# 데이터 모델
# ============================================================

@dataclass
class ParsedRow:
    """파싱된 단일 행 데이터."""
    reporting_starts: str = ""
    reporting_ends: str = ""
    ad_name: str = ""
    placement: str = ""
    age: str = ""
    gender: str = ""
    amount_spent: float = 0.0
    link_clicks: int = 0
    ctr: float = 0.0
    cpc: float = 0.0
    purchases: int = 0
    purchase_conversion_value: float = 0.0
    purchase_roas: float = 0.0


@dataclass
class ParsedReport:
    """단일 .xlsx 파일 파싱 결과."""
    rows: list[ParsedRow] = field(default_factory=list)
    column_map: dict[str, str] = field(default_factory=dict)  # 표준키 → 실제 헤더명
    missing_optional: list[str] = field(default_factory=list)  # 누락 선택 컬럼
    has_demographics: bool = True  # 연령·성별 모두 있으면 True


@dataclass
class ValidationResult:
    """정합성 검증 결과 (6 체크, SPEC §3.4 부록 H.1)."""
    # 체크 1: 파일 수 범위 (1~6개)
    file_count_valid: bool = True
    file_count: int = 0
    # 체크 2: 지출 합산 일치
    spend_consistent: bool = True
    spend_total: float = 0.0
    # 체크 3: 노출 합산 일치 (있다면)
    impressions_consistent: bool | None = None  # None = 데이터 없음
    # 체크 4: 구매수 일치
    purchases_consistent: bool = True
    purchases_total: int = 0
    # 체크 5: 기간 라벨 일치
    period_consistent: bool = True
    period_labels: list[str] = field(default_factory=list)
    # 체크 6: 통합/분리 판정
    report_type: str = "unknown"  # "integrated" | "segmented" | "unknown"
    # 전체 통과 여부
    overall_valid: bool = True
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class XlsxParseOutput:
    """xlsx 파싱 전체 출력."""
    reports: list[ParsedReport]
    validation: ValidationResult
    # 통합 데이터 (여러 파일 합산)
    merged_rows: list[ParsedRow] = field(default_factory=list)
    # 블로커 리포트 (누락 처리 룰 3 — 자동 매핑 실패 시)
    blocker: str | None = None


# ============================================================
# 헬퍼 함수
# ============================================================

def _normalize_header(raw: str) -> str:
    """헤더 정규화 — 공백·대소문자 무시."""
    return raw.strip().lower().replace(" ", "")


def _build_column_map(headers: list[str]) -> tuple[dict[str, str], list[str], list[str]]:
    """
    헤더 리스트 → 표준 컬럼 키 매핑.
    누락 필수 컬럼과 누락 선택 컬럼을 분리 반환.

    Returns:
        (column_map, missing_mandatory, missing_optional)
    """
    normalized_headers = {_normalize_header(h): h for h in headers}
    column_map: dict[str, str] = {}
    missing_mandatory: list[str] = []
    missing_optional: list[str] = []

    for std_key, aliases in REQUIRED_COLUMN_ALIASES.items():
        found = False
        for alias in aliases:
            norm = _normalize_header(alias)
            if norm in normalized_headers:
                column_map[std_key] = normalized_headers[norm]
                found = True
                break
        if not found:
            if std_key in MANDATORY_COLUMNS:
                missing_mandatory.append(std_key)
            else:
                missing_optional.append(std_key)

    return column_map, missing_mandatory, missing_optional


def _safe_float(val: Any) -> float:
    """안전한 float 변환. 변환 실패 시 0.0."""
    if val is None or val == "":
        return 0.0
    try:
        return float(str(val).replace(",", "").replace("₩", "").replace("%", ""))
    except (ValueError, TypeError):
        return 0.0


def _safe_int(val: Any) -> int:
    """안전한 int 변환. 변환 실패 시 0."""
    f = _safe_float(val)
    return int(f)


def _safe_str(val: Any) -> str:
    """안전한 str 변환."""
    if val is None:
        return ""
    return str(val).strip()


# ============================================================
# 핵심 파서
# ============================================================

def _parse_workbook(wb: Workbook) -> ParsedReport:
    """
    openpyxl Workbook → ParsedReport.
    PIPA: 메모리 내 처리만, 디스크 저장 없음.
    """
    ws = wb.active
    if ws is None:
        return ParsedReport()

    # 헤더 행 추출 (첫 번째 비빈 행)
    headers: list[str] = []
    header_row_idx = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        non_empty = [c for c in row if c is not None and str(c).strip()]
        if non_empty:
            headers = [str(c).strip() if c is not None else "" for c in row]
            header_row_idx = i
            break

    if not headers:
        return ParsedReport()

    column_map, missing_mandatory, missing_optional = _build_column_map(headers)

    # 필수 컬럼 누락 룰 1 (SPEC §3.4 누락 처리 4룰):
    # 필수 컬럼 누락 → 명확한 오류 (blocker는 XlsxParseOutput 레벨에서 처리)
    parsed_report = ParsedReport(
        column_map=column_map,
        missing_optional=missing_optional,
    )

    # 연령·성별 누락 룰 2 — 인구통계 분석 비활성
    if "age" in missing_optional or "gender" in missing_optional:
        parsed_report.has_demographics = False

    # 데이터 행 파싱
    rows: list[ParsedRow] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(c is None for c in row):
            continue  # 빈 행 스킵

        def get(std_key: str) -> Any:
            """표준 키로 현재 행 값 추출."""
            actual_header = column_map.get(std_key)
            if actual_header is None:
                return None
            try:
                col_idx = headers.index(actual_header)
                return row[col_idx]
            except (ValueError, IndexError):
                return None

        # CTR 누락 룰 4 — CTR 기반 추정 (CTR=0 제외)
        ctr_val = _safe_float(get("ctr"))
        link_clicks_val = _safe_int(get("link_clicks"))
        if link_clicks_val == 0 and ctr_val > 0:
            # 노출수 누락 시 CTR 기반 추정 (CTR=0 제외)
            # link_clicks는 실측값 우선, 없으면 0 유지
            pass

        parsed_row = ParsedRow(
            reporting_starts=_safe_str(get("reporting_starts")),
            reporting_ends=_safe_str(get("reporting_ends")),
            ad_name=_safe_str(get("ad_name")),
            placement=_safe_str(get("placement")),
            age=_safe_str(get("age")),
            gender=_safe_str(get("gender")),
            amount_spent=_safe_float(get("amount_spent")),
            link_clicks=link_clicks_val,
            ctr=ctr_val,
            cpc=_safe_float(get("cpc")),
            purchases=_safe_int(get("purchases")),
            purchase_conversion_value=_safe_float(get("purchase_conversion_value")),
            purchase_roas=_safe_float(get("purchase_roas")),
        )
        rows.append(parsed_row)

    parsed_report.rows = rows
    return parsed_report


def _validate_reports(reports: list[ParsedReport], file_count: int) -> ValidationResult:
    """
    정합성 검증 6 체크 — SPEC §3.4 부록 H.1.

    체크 1: 파일 수 범위 (1~6개)
    체크 2: 지출 합산 일치
    체크 3: 노출 합산 일치 (있다면)
    체크 4: 구매수 일치
    체크 5: 기간 라벨 일치
    체크 6: 통합/분리 판정
    """
    result = ValidationResult(file_count=file_count)
    errors: list[str] = []
    warnings: list[str] = []

    # 체크 1: 파일 수 1~6개
    if not (1 <= file_count <= 6):
        result.file_count_valid = False
        errors.append(f"파일 수 범위 초과: {file_count}개 (허용: 1~6개)")

    # 통합 집계
    all_rows = [r for report in reports for r in report.rows]
    if not all_rows:
        result.overall_valid = False
        result.errors = errors
        result.warnings = warnings
        return result

    total_spend = sum(r.amount_spent for r in all_rows)
    total_purchases = sum(r.purchases for r in all_rows)
    result.spend_total = round(total_spend, 2)
    result.purchases_total = total_purchases

    # 체크 2: 지출 합산 — 파일 수 > 1이면 분리 보고서 합산 일치 검사
    if file_count > 1:
        per_file_spends = [sum(r.amount_spent for r in report.rows) for report in reports]
        # 분리 보고서에서 각 파일 지출합 합산이 전체와 일치해야 함 (허용 오차 1%)
        combined = sum(per_file_spends)
        if combined > 0 and abs(combined - total_spend) / combined > 0.01:
            result.spend_consistent = False
            warnings.append("지출 합산 불일치: 파일 간 지출 합계 불일치 가능성 검토 권장")

    # 체크 3: 노출 합산 (link_clicks가 있는 경우)
    total_clicks = sum(r.link_clicks for r in all_rows)
    if total_clicks > 0:
        result.impressions_consistent = True  # link_clicks 기반 간접 검증

    # 체크 4: 구매수 일치 — 구매 ROAS와 전환값으로 교차 검증
    if total_spend > 0 and total_purchases > 0:
        implied_roas = (
            sum(r.purchase_conversion_value for r in all_rows) / total_spend
        )
        avg_roas = sum(r.purchase_roas for r in all_rows) / len(all_rows)
        if avg_roas > 0 and abs(implied_roas - avg_roas) / avg_roas > 0.15:
            result.purchases_consistent = False
            warnings.append("구매수 일치 검토 권장: 전환값/지출 기반 ROAS와 보고 ROAS 차이 15% 초과")

    # 체크 5: 기간 라벨 일치
    period_labels = list({
        f"{r.reporting_starts}~{r.reporting_ends}"
        for r in all_rows
        if r.reporting_starts or r.reporting_ends
    })
    result.period_labels = period_labels
    if len(period_labels) > 3:
        result.period_consistent = False
        warnings.append(f"기간 라벨 다수 ({len(period_labels)}개): 통합·분리 보고서 혼재 가능성")

    # 체크 6: 통합/분리 판정
    unique_placements = {r.placement for r in all_rows if r.placement}
    unique_ages = {r.age for r in all_rows if r.age}
    unique_genders = {r.gender for r in all_rows if r.gender}
    if len(unique_placements) > 1 or len(unique_ages) > 1 or len(unique_genders) > 1:
        result.report_type = "segmented"  # 분리 보고서 (인구통계 세분화)
    elif file_count > 1:
        result.report_type = "segmented"  # 여러 파일 → 분리 보고서
    else:
        result.report_type = "integrated"  # 단일 파일 통합 보고서

    result.overall_valid = len(errors) == 0
    result.errors = errors
    result.warnings = warnings
    return result


# @MX:ANCHOR: [AUTO] xlsx 파싱 공개 진입점 — fan_in >= 3 (tools 3종 + tests)
# @MX:REASON: REQ-AUDIT-MCP-016 "12 필수 컬럼 자동 매핑 + 정합성 6 체크 + 누락 4 룰" 공식 인터페이스.
#             REQ-AUDIT-MCP-019 PIPA 메모리 처리 보장. 이 함수가 유일한 외부 진입점.
def parse_xlsx_files(xlsx_paths: list[str]) -> XlsxParseOutput:
    """
    .xlsx 파일 1~6개 파싱 + 정합성 검증.

    PIPA 준수: 파일 내용은 메모리에서만 처리, 중간 가공물 디스크 저장 없음.
    REQ-AUDIT-MCP-019: openpyxl 메모리 스트림 방식 사용.

    누락 처리 4룰 (SPEC §3.4 부록 B.5):
    룰 1: 필수 컬럼 누락 → 명확한 오류 (blocker 설정)
    룰 2: 연령·성별 누락 → has_demographics=False (인구통계 분석 비활성)
    룰 3: 자동 매핑 실패 → blocker 리포트 (사용자 확인 요청)
    룰 4: 노출수 누락 → CTR 기반 추정 (CTR=0 제외)

    Args:
        xlsx_paths: .xlsx 파일 절대 경로 리스트 (1~6개)

    Returns:
        XlsxParseOutput (reports, validation, merged_rows, blocker)
    """
    reports: list[ParsedReport] = []
    blockers: list[str] = []

    for path in xlsx_paths:
        if not os.path.isfile(path):
            blockers.append(f"파일 없음: {path}")
            continue

        # PIPA: read_only=True + data_only=True (수식 값만, 메모리 최소화)
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            blockers.append(f"파일 파싱 오류: {os.path.basename(path)} — {exc}")
            continue

        report = _parse_workbook(wb)
        wb.close()  # 스트림 즉시 해제 (PIPA)

        # 룰 1: 필수 컬럼 누락 → 블로커
        missing_mandatory = [
            k for k in MANDATORY_COLUMNS if k not in report.column_map
        ]
        if missing_mandatory:
            blockers.append(
                f"필수 컬럼 누락 ({os.path.basename(path)}): "
                f"{', '.join(missing_mandatory)}"
            )
            continue

        # 룰 3: 자동 매핑 실패 컬럼이 있으면 경고 (블로커 아님 — 선택 컬럼)
        if report.missing_optional:
            # 선택 컬럼 누락은 blocker 아닌 warning (룰 3 — 블로커는 매핑 자체 불가 시)
            pass

        reports.append(report)

    # 정합성 검증 (블로커 없이 파싱 성공한 파일만)
    validation = _validate_reports(reports, file_count=len(xlsx_paths))

    # 병합 데이터
    merged_rows = [r for report in reports for r in report.rows]

    # 최종 블로커 결정
    blocker_msg: str | None = None
    if blockers:
        blocker_msg = "\n".join(blockers)

    return XlsxParseOutput(
        reports=reports,
        validation=validation,
        merged_rows=merged_rows,
        blocker=blocker_msg,
    )


def parse_xlsx_bytes(xlsx_bytes: bytes, filename: str = "report.xlsx") -> XlsxParseOutput:
    """
    바이트 데이터로 단일 .xlsx 파싱 (Layer 1 MCP 출력 수신 시 사용).
    PIPA: 디스크 저장 없음, 메모리 스트림만 사용.

    Args:
        xlsx_bytes: .xlsx 파일 바이트 데이터
        filename: 오류 메시지용 파일명 (경로 아님)

    Returns:
        XlsxParseOutput
    """
    try:
        stream = io.BytesIO(xlsx_bytes)
        wb = openpyxl.load_workbook(stream, read_only=True, data_only=True)
        report = _parse_workbook(wb)
        wb.close()
    except Exception as exc:
        return XlsxParseOutput(
            reports=[],
            validation=ValidationResult(file_count=0, overall_valid=False),
            blocker=f"바이트 파싱 오류: {filename} — {exc}",
        )

    validation = _validate_reports([report], file_count=1)
    return XlsxParseOutput(
        reports=[report],
        validation=validation,
        merged_rows=report.rows,
        blocker=None,
    )
